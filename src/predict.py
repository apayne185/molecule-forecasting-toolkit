from pathlib import Path

import numpy as np
import pandas as pd
import openpyxl

from .features import FEATURES, CAT_FEATURES, START_ROW_EXCEL, MAX_FORECAST_ROWS


def generate_forecasts(
    model_value,
    model_packs,
    future_df: pd.DataFrame,
) -> pd.DataFrame:
    """Run both models on future_df and return a DataFrame with predictions attached."""
    future_X = future_df[FEATURES].copy()
    for col in CAT_FEATURES:
        future_X[col] = future_X[col].astype('category')

    result = future_df.copy()
    result['Predicted_Value'] = model_value.predict(future_X)
    result['Predicted_Packs'] = model_packs.predict(future_X)
    return result


def save_to_excel(
    predictions: pd.DataFrame,
    template_path: str | Path,
    output_path: str | Path,
) -> None:
    """Write Value and Packs predictions into the Excel template at the target range."""
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    values = predictions['Predicted_Value'].tolist()[:MAX_FORECAST_ROWS]
    packs  = predictions['Predicted_Packs'].tolist()[:MAX_FORECAST_ROWS]

    if len(values) != len(packs):
        raise ValueError(
            f"Length mismatch: {len(values)} Value predictions vs {len(packs)} Packs predictions"
        )

    for i, (val, pack) in enumerate(zip(values, packs), start=START_ROW_EXCEL):
        ws[f'G{i}'] = val
        ws[f'H{i}'] = pack

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Saved {len(values)} forecasts → {output_path}")
